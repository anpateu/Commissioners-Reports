import openpyxl
import os
from config import Config
from api_ozon import OzonAPI

def create_Excel_for_1C(response):
    offer_ids = []
    returns_amount = 0
    ozon = openpyxl.Workbook()
    ozon.create_sheet(title='OZON', index=0)
    sheet_ozon = ozon['OZON']
    sheet_ozon.append(['Код', 'Наименование', 'Кол-во', 'Цена', 'Сумма'])

    returns_report = openpyxl.Workbook()
    returns_report.create_sheet(title='Returns', index=0)
    sheet_returns = returns_report['Returns']
    sheet_returns.append(['Код', 'Наименование', 'Кол-во', 'Цена', 'Сумма'])

    for element in response.json()['result']['rows']:
        offer_id = element['offer_id']
        product_name = element['product_name']
        amount = 0.0
        price = 0.0
        summary = 0.0

        if offer_id not in offer_ids:
            for offer_element in response.json()['result']['rows']:
                if offer_element['offer_id'] == offer_id:
                    amount += offer_element['sale_qty']
                    amount -= offer_element['return_qty']
                    summary += offer_element['sale_amount']
                    summary += offer_element['sale_discount']
                    summary -= offer_element['return_amount']
                    summary -= offer_element['return_discount']

            if amount != 0:
                price = summary / amount

            amount = round(amount, 2)
            summary = round(summary, 2)

            offer_ids.append(offer_id)
            if amount != 0:
                if amount > 0:
                    sheet_ozon.append([offer_id, product_name, amount, price, summary])
                else:
                    returns_amount += 1
                    sheet_returns.append([offer_id, product_name, amount, price, summary])

    number = response.json()['result']['header']['num']
    doc_date = response.json()['result']['header']['doc_date']
    ozon.save(f"Ozon\\НП 1С Отчет комиссионера ОЗОН от {doc_date} №{number}.xlsx")
    print(f'Created report: date {doc_date} - N{number}')
    if returns_amount > 0:
        returns_report.save(f"Ozon\\НП 1С Возвраты ОЗОН от {doc_date} №{number}.xlsx")
        print(f'> total items returned: {returns_amount}')


if __name__ == '__main__':
    os.makedirs("Ozon", exist_ok=True)
    os.makedirs("Wildberries", exist_ok=True)

    api = OzonAPI(client_id=Config.OZON_CLIENT_ID, api_key=Config.OZON_API_KEY)
    create_Excel_for_1C(api.get_report())