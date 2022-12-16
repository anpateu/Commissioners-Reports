import calendar
from datetime import datetime, timedelta
from api_wb import WildberriesAPI
from api_ms import MoySkladAPI
from config import Config
import time
import os
import openpyxl


def load_error_codes():
    with open('data/error_codes.txt', 'r', encoding='utf-8') as f:
        dict = { }
        contents = f.read()
        lines = contents.split('\n')
        for line in lines:
            key, value = line.split(':')
            dict[key] = value
    return dict

def check_articul(articul, error_codes):
    if articul in error_codes:
        articul = error_codes[articul]
    return articul

def check_dates():
    year = datetime.now().year
    month = datetime.now().month
    first_day, last_day = calendar.monthrange(year, month - 1)

    monday_dates = []
    for day in range(first_day, last_day + 1):
        date = datetime(year, month - 1, day)
        if date.weekday() == calendar.MONDAY:
            monday_dates.append(date)

    if monday_dates[0].strftime("%d") != '01':
        previous_week = monday_dates[0] - timedelta(weeks=1)
        monday_dates.insert(0, previous_week)

    start_dates = []
    end_dates = []

    for date in monday_dates:
        delta = date + timedelta(days=6)
        if delta.month != month:
            start_dates.append(date.strftime("%Y-%m-%dT00:00:00"))
            end_dates.append(delta.strftime("%Y-%m-%dT23:59:59"))

    return start_dates, end_dates

def create_Excel_for_1C(data, start_date, end_date, error_codes):
    returns_amount = 0
    offer_ids = []
    wb = openpyxl.Workbook()
    wb.create_sheet(title='WB', index=0)
    sheet_wb = wb['WB']
    sheet_wb.append(['Код', 'Кол-во', 'Цена', 'Сумма'])

    returns_report = openpyxl.Workbook()
    returns_report.create_sheet(title='Returns', index=0)
    sheet_returns = returns_report['Returns']
    sheet_returns.append(['Код', 'Кол-во', 'Сумма', 'Дата продажи', 'Дата возврата'])

    for element in data:
        offer_id = check_articul(element['sa_name'], error_codes)
        type = element['supplier_oper_name']
        amount = 0.0
        price = 0.0
        summary = 0.0

        if offer_id not in offer_ids and element['quantity'] != 0:
            for offer_element in data:
                offer_element_id = check_articul(offer_element['sa_name'], error_codes)
                if offer_element_id == offer_id and offer_element['quantity'] != 0:
                    country = offer_element['site_country']
                    if country == 'RU' or country == '':
                        type = offer_element['supplier_oper_name']
                        if type == 'Продажа' or type == 'Оплата брака':
                            amount += offer_element['quantity']
                            summary += offer_element['retail_amount']
                        else:
                            if type != 'Возврат':
                                print(type, offer_element)
                    else:
                        if country != 'RU' and country != '' and country != 'BY':
                            print(country, offer_element)

            if amount != 0:
                price = summary / amount

            amount = round(amount, 2)
            summary = round(summary, 2)

            offer_ids.append(offer_id)
            if amount != 0:
                if amount > 0:
                    sheet_wb.append([offer_id, amount, price, summary])

        if type == 'Возврат':
            amount = element['quantity']
            summary = element['retail_amount']
            sale_date = datetime.strptime(element['order_dt'], "%Y-%m-%dT%H:%M:%SZ").strftime("%Y-%m-%d")
            return_date = datetime.strptime(element['sale_dt'], "%Y-%m-%dT%H:%M:%SZ").strftime("%Y-%m-%d")
            sheet_returns.append([offer_id, amount, summary, sale_date, return_date])
            returns_amount += 1

    start = datetime.strptime(start_date, "%Y-%m-%dT%H:%M:%S").strftime("%Y-%m-%d")
    end = datetime.strptime(end_date, "%Y-%m-%dT%H:%M:%S").strftime("%Y-%m-%d")
    wb.save(f"Wildberries\\НП 1С Отчет комиссионера ВБ с {start} по {end}.xlsx")
    print(f'Created report')
    if returns_amount > 0:
        returns_report.save(f"Wildberries\\НП 1С Возвраты ВБ c {start} по {end}.xlsx")
        print(f'> total items returned: {returns_amount}')


if __name__ == '__main__':
    os.makedirs("Ozon", exist_ok=True)
    os.makedirs("Wildberries", exist_ok=True)

    api = WildberriesAPI(api_key=Config.WB_API_KEY)
    api_ms = MoySkladAPI(api_key=Config.MS_API_KEY)

    start_dates, end_dates = check_dates()
    error_codes = load_error_codes()

    for k in range(len(start_dates)):
        start_date = start_dates[k]
        end_date = end_dates[k]
        create_Excel_for_1C(api.get_report(start_date, end_date), start_date, end_date, error_codes)
        time.sleep(60)